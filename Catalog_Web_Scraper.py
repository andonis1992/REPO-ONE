from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
import os, sys #Standard python lib
import time, datetime




#list of product numbers
product_list = []
#list of vapor pressures
vapor_list = []


os.chdir(sys.path[0])
#loading excel product numbers
WB1 = load_workbook("VaporPressures.xlsx", data_only = True)
WB1_WS1 = WB1["Sheet1"]


p_index = 2

while WB1_WS1.cell(column = 1, row = p_index).value != None:
    product_list.append(WB1_WS1.cell(column = 1, row = p_index).value)
    p_index +=1


#Function to find Vapor pressure and append vapor pressure list
def find_vapor(prod_num, list):
    
    url_list = ['https://www.sigmaaldrich.com/US/en/product/sigald/'+ str(prod_num), 'https://www.sigmaaldrich.com/US/en/product/aldrich/'+ str(prod_num), "https://www.sigmaaldrich.com/US/en/product/sial/"+ str(prod_num), "https://www.sigmaaldrich.com/US/en/product/aldrich/"+ str(prod_num) +"?context=product"]
    # Start the browser and navigate to the URL
    service = Service("C:\\Users\\M310648\\.spyder-py3\\Chrome DONT DELTE OR MOVE\\chromedrier.exe")
    options = webdriver.Chrome(service=service)
    

    #vapor_pressure = options.find_element(By.XPATH, '//*[@id="prodductDetailGrid"]/div[3]/div/div[2]/div/div/div/div/div/div/div[1]/div/div[6]/div[2]/p/span')
    #print(vapor_pressure.text)

    
    #Searches through all text to find the word vapor pressure, then chooses the next item on the list which is the actual value
    
    index = 0 
    url_index = 0   
    while url_index < 4:
        options.get(str(url_list[url_index])) 
        text_elements = options.find_element(By.TAG_NAME, 'body').text
        text_list = text_elements.split("\n") 
        #print(text_list)
        text_index = 0
        string = "vapor pressure"
        if url_index > 4:
            continue
        else:
            for items in text_list:                
                if items == string:
                    list.append(str(text_list[text_index+1]))
                    index +=1
                    url_index = 25
                    continue            
                else:                
                    index += 1
                    text_index += 1                    
        url_index += 1
        if url_index == 3:
            list.append("None")
        else:
            continue
        
#Calling finding vapor function
for items in product_list:
    find_vapor(items, vapor_list)
    print(vapor_list)


print(vapor_list)




workbook = Workbook()
Export = workbook.active
Export.title = "Export"

#Creation of column headers
Export["A1"] = "Product Number"
Export["B1"] = "Vapor Pressure"

#defining function to export a column list to a column in excel

def export_to_excel(column_list, Letter):
    i = 2
    for strings in column_list:
        Export[Letter + str(i)] = strings
        i +=1
export_to_excel(product_list, "A")
export_to_excel(vapor_list, "B")


t = time.localtime()
timestamp = time.strftime('%b-%d-%Y_%H%M', t)
BACKUP_NAME = ("Vapor Pressures-" + timestamp)
workbook.save(filename=BACKUP_NAME +".xlsx")