from asyncio.windows_events import NULL
from itertools import product
from math import prod
import os, sys #Standard python lib
from docxtpl import DocxTemplate
import time, datetime
from openpyxl import Workbook, load_workbook
import pandas as pd

#Change path to current working directory
os.chdir(sys.path[0])

#Creating Timestamp for initial release dating
t = time.localtime()
timestamp = time.strftime('%b-%d-%Y_%H%M', t)
datestamp = time.strftime('%b-%d-%Y')



# Loading Data from excel
WB1 = load_workbook("Protocol_Update_Template.xlsx", data_only = True)
WB1_WS1 = WB1["Export"]

#Creating Timestamp that can be adjusted based on expiry
initial_date = time.strftime('%b-14-%Y')

def timestampcalc (year_num, month_num):
    req_date = pd.to_datetime(initial_date) + pd.DateOffset(years=year_num) + pd.DateOffset(months=month_num)
    req_date = req_date.strftime('%m-%d-%Y')
    return req_date



#Defining lists and variables
product_family = []
product_num = []
product_name = []
batch_num = []
temp = []
temp_index = []
expiry = []
year_one = []
year_one_text = []
year_two = []
year_two_text = []
year_three = []
year_three_text = []
year_four = []
year_four_text = []
year_five = []
year_five_text = []
shelf_life = []
date_initial = []
total_samples = []
testing_strings = []
test_one = []
test_two = []
test_three = []
test_four = []
test_one_doc = []
test_two_doc = []
test_three_doc = []
test_four_doc = []
forcast_index = 2
expiry_index = 0

#Searching through excel to find values for each product
while WB1_WS1.cell(column = 1, row = forcast_index).value != None:
    product_family.append(WB1_WS1.cell(column = 5, row = forcast_index).value)
    product_num.append(WB1_WS1.cell(column = 1, row = forcast_index).value)
    batch_num.append(WB1_WS1.cell(column = 2, row = forcast_index).value)
    expiry.append(WB1_WS1.cell(column = 8, row = forcast_index).value)
    product_name.append(WB1_WS1.cell(column = 3, row = forcast_index).value)
    temp_index.append(WB1_WS1.cell(column = 9, row = forcast_index).value)
    testing_strings.append(WB1_WS1.cell(column = 20, row = forcast_index).value)
    date_initial.append(datestamp)
    
    #Findind the expiry, and associating values to each corresponding list based on length of expiry
    if expiry[expiry_index] == '1825' or expiry[expiry_index] == 1825:
        shelf_life.append('5 Y')
        year_one.append(timestampcalc(1, 0))
        year_one_text.append('1 Year')
        year_two.append(timestampcalc(2, 0))
        year_two_text.append('2 Year')
        year_three.append(timestampcalc(3, 0))
        year_three_text.append('3 Year')
        year_four.append(timestampcalc(4, 0))
        year_four_text.append('4 Year')
        year_five.append(timestampcalc(5, 0))
        year_five_text.append('5 Year')
        total_samples.append('7')
    elif expiry[expiry_index] == '912' or expiry[expiry_index] == 912:
        shelf_life.append('2.5 Y')
        year_one.append(timestampcalc(1, 0))
        year_one_text.append('1 Year')
        year_two.append(timestampcalc(1, 6))
        year_two_text.append('1.5 Year')
        year_three.append(timestampcalc(2, 0))
        year_three_text.append('2 Year')
        year_four.append(timestampcalc(2, 6))
        year_four_text.append('2.5 Year')
        year_five.append(" ")
        year_five_text.append(" ")
        total_samples.append('6')
    elif expiry[expiry_index] == '1095' or expiry[expiry_index] == 1095:
        shelf_life.append('3 Y')
        year_one.append(timestampcalc(1, 0))
        year_one_text.append('1 Year')
        year_two.append(timestampcalc(1, 6))
        year_two_text.append('1.5 Year')
        year_three.append(timestampcalc(2, 0))
        year_three_text.append('2 Year')
        year_four.append(timestampcalc(2, 6))
        year_four_text.append('2.5 Year')
        year_five.append(timestampcalc(3, 0))
        year_five_text.append('3 Year')
        total_samples.append('7')
        
    #associating the correct temp, based on storage condition
    if temp_index[expiry_index] == 'RT':
        temp.append('25C')
    elif temp_index[expiry_index] == 'CL':
        temp.append('5C')
    forcast_index += 1
    expiry_index += 1

#Print Check uncomment for troubleshooting
# print(product_family)
# print(product_name)
# print(batch_num)
# print(str(product_num)[0:7])
# print(expiry)
# print(year_one)
# print(year_two)
# print(year_three)
# print(year_four)
for items in testing_strings:
    try:
        test_one.append(items.split(",")[0])
    except:
        test_one.append(" ")
    try:
        test_two.append(items.split(",")[1])
    except: 
        test_two.append(" ")
    try:
        test_three.append(items.split(",")[2])
    except:
        test_two.append(" ")
    try:
        test_four.append(items.split(",")[3])
    except:
        test_four.append(" ")
    
def find_doc(list, doc_list):
    for item in list:
        if 'ATR' in item:
            doc_list.append("15123832-LS")
        elif "APP" in item:
            doc_list.append("17894080-LS")
        elif "VPCC" in item:
            doc_list.append("15070820-LS")
        elif "BHT" in item:
            doc_list.append("23862366-LS")
        elif "VPCT" in item:
            doc_list.append("15070820-LS")
        elif "GC" in item:
            doc_list.append("15070820-LS")
        elif "IRA" in item:
            doc_list.append("15355233-LS")
        elif "Stabilizer" in item:
            doc_list.append("23862366-LS")
        elif "CLO4" in item or "NaOH" in item:
            doc_list.append("20572958-LS")
        elif "NMR" in item:
            doc_list.append("15355233-LS")
        elif "MP" in item:
            doc_list.append("15069595-LS")            
        
        elif item == " ":
            doc_list.append(" ")
        else:
            doc_list.append("document number not found")
find_doc(test_one, test_one_doc)  
find_doc(test_two, test_two_doc) 
find_doc(test_three, test_three_doc) 
find_doc(test_four, test_four_doc) 
            
# print(testing_strings)
# print(test_one)
# print(test_two)
# print(test_three)
# print(test_four)
# print(test_one_doc)
# print(test_two_doc)
# print(test_three_doc)
# print(test_four_doc)


#Looping through lists and creating a word document for each product
list_index = 0
while forcast_index > 2:
    BACKUP_NAME = ('Shelf Life Protocol '+ str(product_num[list_index]))
    doc = DocxTemplate('Template.docx')

    context = {'family_upper': product_family[list_index].upper(), 'product_family': product_family[list_index], 'batch_num': batch_num[list_index], 'product_num': product_num[list_index][0:7],\
        'year_one': str(year_one[list_index]), 'year_two': year_two[list_index], 'year_three': year_three[list_index], 'year_four': year_four[list_index], 'year_five': year_five[list_index],\
            'product_name': product_name[list_index], 'shelf_life': shelf_life[list_index], 'Temp': temp[list_index], 'date_initial': date_initial[list_index], 'year_one_text': year_one_text[list_index], \
                'year_two_text': year_two_text[list_index], 'year_three_text': year_three_text[list_index], 'year_four_text': year_four_text[list_index], 'year_five_text': year_five_text[list_index], \
                    'total_samples': total_samples[list_index], 'test_one': test_one[list_index], 'test_two': test_two[list_index], 'test_three': test_three[list_index], 'test_four': test_four[list_index],\
                     'test_one_doc': test_one_doc[list_index], 'test_two_doc': test_two_doc[list_index], 'test_three_doc': test_three_doc[list_index], 'test_four_doc': test_four_doc[list_index] }

    doc.render(context)

    doc.save(BACKUP_NAME+'.docx')
    list_index += 1
    forcast_index -= 1
    