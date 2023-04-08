from asyncio.windows_events import NULL
from itertools import product
import os, sys #Standard python lib
from docxtpl import DocxTemplate
import time, datetime
from openpyxl import Workbook, load_workbook
from tkinter import * 

#Change path to current working directory
os.chdir(sys.path[0])




def scarautomation(file_name):
    # Loading Data from excel
    WB1 = load_workbook(file_name+".xlsx", data_only = True)
    WB1_WS1 = WB1["Sheet1"]

    #Creating Timestamp 
    t = time.localtime()
    timestamp = time.strftime('%b-%d-%Y_%H%M', t)
    datestamp = time.strftime('%b-%d-%Y')

    #Defining Variables
    complaint_num = WB1_WS1.cell(column = 15, row = 4).value
    contact_address = str
    global_item_number = str
    reference_number = str
    product_name = str
    lot_number = str
    invest_full = str
    description_one = str
    #Finding and Copying Contact Address
    row_index = 1
    while row_index <= 1000:
        if WB1_WS1.cell(column = 2, row = row_index).value == "Contact Address:":
            contact_address = WB1_WS1.cell(column = 16, row = row_index).value
            break
        else:
            row_index += 1
            
    #Finding and Copying Global Item Number
    row_index = 1
    while row_index <= 1000:
        if WB1_WS1.cell(column = 2, row = row_index).value == "Global Item Number:":
            global_item_number = WB1_WS1.cell(column = 16, row = row_index).value
            break
        else:
            row_index += 1
            
            
    #Finding and Copying Reference Number
    row_index = 1
    while row_index <= 1000:
        if WB1_WS1.cell(column = 29, row = row_index).value == "Customer Reference:":
            reference_number = WB1_WS1.cell(column = 41, row = row_index).value
            if reference_number == "- - -":
                reference_number = 'N/A'
            break
        else:
            row_index += 1
            
    #Finding and Product Name
    row_index = 1
    while row_index <= 1000:
        if WB1_WS1.cell(column = 2, row = row_index).value == "Material Name:":
            product_name = WB1_WS1.cell(column = 16, row = row_index).value
            break
        else:
            row_index += 1
            
    #Finding and Lot Number
    row_index = 1
    while row_index <= 1000:
        if WB1_WS1.cell(column = 2, row = row_index).value == "Batch/Lot Number:":
            lot_number = WB1_WS1.cell(column = 16, row = row_index).value
            break
        else:
            row_index += 1
                    
    #Finding and Investigation Results
    row_index = 1
    while row_index <= 1000:
        if WB1_WS1.cell(column = 2, row = row_index).value == "Investigation Results:":
            invest_full = WB1_WS1.cell(column = 16, row = row_index).value
            break
        else:
            row_index += 1
            
    #Finding and Investigation Results
    row_index = 1
    while row_index <= 1000:
        if WB1_WS1.cell(column = 2, row = row_index).value == "Description:":
            description_one = WB1_WS1.cell(column = 16, row = row_index).value
            break
        else:
            row_index += 1
    #Print Check
    print(contact_address)
    print(complaint_num)
    print(str(global_item_number)[0:5])
    print(reference_number)
    print(product_name)
    print(lot_number)
    print(invest_full)
    opening_statement = "You reported an issue with " + product_name + ". Please see below for the final report of the investigation we have performed."

    invest_parts = invest_full.split('\n')
    print(invest_parts)
    BACKUP_NAME = ('TW '+ str(complaint_num) + " " + str(global_item_number)[0:5] + " SCAR" )
    doc = DocxTemplate('Template.docx')

    context = {'Date': datestamp, 'contact_address': contact_address, 'complaint_num': complaint_num, 'global_item_number': global_item_number, 'reference_number': reference_number, 'Product_name': product_name,\
        'Lot_Number': lot_number, 'opening_statement': opening_statement, 'investigation_full': invest_full, 'Description': description_one, 'global_item_number_short': global_item_number[0:5] }

    doc.render(context)

    doc.save(BACKUP_NAME+'.docx')

#Must always go first creates window screen
root = Tk()

#Defining input box
e = Entry(root, width =35)



#Defining click function
def myClick():
    myLabel3 = Label(root, text="Please Wait, Generating Report " + e.get())
    myLabel3.grid(row=5, column=0)
    scarautomation(e.get())
    myLabel3.grid_remove()
    myLabel6 = Label(root, text="Report Complete")
    myLabel6.grid(row=5, column=0)

#Creating label widget
#Define the label
myLabel1 = Label(root, text = 'SCAR Generator')
myLabel2= Label(root, text = 'Please enter a TW Excel')
myLabel4= Label(root, text="Only enter file name")
#define Button
myButton = Button(root, text='Run', command=myClick, fg="blue", bg="yellow")
#Putting the label on the screen
myLabel1.grid(row=0, column=0)
myLabel2.grid(row=1, column=0)
myLabel4.grid(row=2, column=0)
myButton.grid(row=4, column=0)
e.grid(row=3, column=0)


root.mainloop()